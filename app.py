# --- Per-user DB Management ---
def get_user_db_list(user_id):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS user_db (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        db_name TEXT NOT NULL,
        UNIQUE(user_id, db_name)
    )''')
    c.execute('SELECT db_name FROM user_db WHERE user_id=?', (user_id,))
    dbs = [row[0] for row in c.fetchall()]
    conn.close()
    return dbs

def get_db_path(user_id, db_name):
    return f"people_{user_id}_{db_name}.db"
import json
import hashlib
from flask import Flask, render_template, request, redirect, url_for, make_response
import sqlite3
import os
import uuid
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB limit

# User table for login
def init_user_db():
    if not os.path.exists('users.db'):
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        c.execute('''
            CREATE TABLE user (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()
init_user_db()


# --- User DB Management ---
from flask import session
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'devkey')

def get_logged_in_user():
    user_id = session.get('user_id')
    if not user_id:
        return None
    return user_id

def init_db(db_name):
    if not os.path.exists(db_name):
        conn = sqlite3.connect(db_name)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE person (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                forename TEXT NOT NULL,
                middle_name TEXT,
                surname TEXT NOT NULL,
                dob TEXT NOT NULL,
                subject TEXT NOT NULL,
                grade INTEGER NOT NULL CHECK(grade >= 0 AND grade <= 100)
            )
        ''')
        conn.commit()
        conn.close()
    # Create upload folder if not exists
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])

init_db("people.db")

## Duplicate cookie_consent route removed

## Duplicate index route removed

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        if not username or not password:
            error = "Username and password required."
        else:
            conn = sqlite3.connect('users.db')
            c = conn.cursor()
            c.execute('SELECT id, password FROM user WHERE username = ?', (username,))
            user = c.fetchone()
            conn.close()
            if user and user[1] == hashlib.sha256(password.encode()).hexdigest():
                resp = make_response(redirect(url_for('index')))
                resp.set_cookie('user_id', str(user[0]), max_age=60*60*24*365)
                return resp
            else:
                error = "Invalid credentials."
    return render_template('login.html', error=error)

@app.route('/graphs', methods=['GET', 'POST'])
def graphs():
    db_name, user_id = get_db_name()
    init_db(db_name)
    axis_options = ['forename', 'middle_name', 'surname', 'dob', 'subject', 'grade']
    graph_type = request.form.get('graph_type', 'line') if request.method == 'POST' else 'line'
    x_axis = request.form.get('x_axis', 'forename') if request.method == 'POST' else 'forename'
    y_axis = request.form.get('y_axis', 'grade') if request.method == 'POST' else 'grade'
    error = None
    # Prevent duplicate axis and require at least one numeric axis
    if x_axis == y_axis:
        error = "X and Y axis must be different."
    if not (x_axis == 'grade' or y_axis == 'grade'):
        error = "At least one axis must be numeric (grade)."
    graph_data = None
    labels, values = [], []
    if not error:
        conn = sqlite3.connect(db_name)
        c = conn.cursor()
        c.execute(f'SELECT {x_axis}, {y_axis} FROM person ORDER BY id')
        rows = c.fetchall()
        conn.close()
        labels = [row[0] for row in rows]
        values = [row[1] for row in rows]
        if labels and values:
            graph_data = {
                'labels': json.dumps(labels),
                'values': json.dumps(values)
            }
    return render_template('graphs.html', graph_type=graph_type, graph_data=graph_data, axis_options=axis_options, x_axis=x_axis, y_axis=y_axis, error=error)
    conn.commit()
    conn.close()
# Create upload folder if not exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

init_db("people.db")



@app.route('/cookie-consent', methods=['GET', 'POST'])
def cookie_consent():
    if request.method == 'POST':
        user_id = str(uuid.uuid4())
        resp = make_response(redirect(url_for('index')))
        resp.set_cookie('user_id', user_id, max_age=60*60*24*365)
        return resp
    return render_template('cookie_consent.html')

@app.route('/')
def index():
    if not request.cookies.get('user_id'):
        return redirect(url_for('cookie_consent'))
    db_name, user_id = get_db_name()
    resp = make_response(render_template('home.html'))
    init_db(db_name)
    return resp

@app.route('/add', methods=['GET', 'POST'])
def add_person():
    db_name, user_id = get_db_name()
    init_db(db_name)
    # Get unique subjects from database for autofill
    conn = sqlite3.connect(db_name)
    c = conn.cursor()
    c.execute('SELECT DISTINCT subject FROM person')
    subjects = [row[0] for row in c.fetchall()]
    conn.close()
    error = None
    if request.method == 'POST':
        forename = request.form['forename'].strip()
        middle_name = request.form.get('middle_name', '').strip()
        surname = request.form['surname'].strip()
        dob = request.form['dob'].strip()
        subject = request.form['subject'].strip()
        grade = request.form['grade'].strip()
        # Data integrity checks
        if not forename or not surname or not dob or not subject or not grade:
            error = "All required fields must be filled."
        try:
            grade_int = int(grade)
            if not (0 <= grade_int <= 100):
                error = "Grade must be between 0 and 100."
        except ValueError:
            error = "Grade must be a number between 0 and 100."
        # Simple date format check (YYYY-MM-DD)
        import re
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", dob):
            error = "Date of Birth must be in YYYY-MM-DD format."
        if not error:
            conn = sqlite3.connect(db_name)
            c = conn.cursor()
            c.execute('''INSERT INTO person (forename, middle_name, surname, dob, subject, grade) VALUES (?, ?, ?, ?, ?, ?)''',
                      (forename, middle_name, surname, dob, subject, grade_int))
            conn.commit()
            conn.close()
            return redirect(url_for('list_people'))
    return render_template('add_person.html', subjects=subjects, error=error)


@app.route('/list', methods=['GET', 'POST'])
def list_people():
    user_id = get_logged_in_user()
    if not user_id:
        return redirect(url_for('login'))
    dbs = get_user_db_list(user_id)
    selected_db = request.args.get('db_name') or (dbs[0] if dbs else None)
    people = []
    if selected_db:
        db_path = get_db_path(user_id, selected_db)
        init_db(db_path)
        # Remove person
        if request.method == 'POST':
            person_id = request.form.get('remove_id')
            if person_id:
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute('DELETE FROM person WHERE id = ?', (person_id,))
                conn.commit()
                conn.close()
        # Search and sort
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'forename')
        order = request.args.get('order', 'asc')
        valid_sort = ['forename', 'middle_name', 'surname', 'dob', 'subject', 'grade']
        if sort_by not in valid_sort:
            sort_by = 'forename'
        order_sql = 'ASC' if order == 'asc' else 'DESC'
        query = 'SELECT id, forename, middle_name, surname, dob, subject, grade FROM person'
        params = []
        if search:
            query += ' WHERE forename LIKE ? OR surname LIKE ?'
            params.extend([f'%{search}%', f'%{search}%'])
        # Special sorting for middle_name: empty values go last
        if sort_by == 'middle_name':
            query += f' ORDER BY CASE WHEN middle_name="" OR middle_name IS NULL THEN 1 ELSE 0 END, middle_name {order_sql}'
        else:
            query += f' ORDER BY {sort_by} {order_sql}'
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute(query, params)
        people = c.fetchall()
        conn.close()
    return render_template('list_people.html', people=people, dbs=dbs, selected_db=selected_db)



# Excel upload route
@app.route('/upload', methods=['GET', 'POST'])
def upload_excel():
    user_id = get_logged_in_user()
    if not user_id:
        return redirect(url_for('login'))
    dbs = get_user_db_list(user_id)
    selected_db = request.form.get('db_name') or (dbs[0] if dbs else None)
    error = None
    success = None
    required_headers = ['forename', 'middle_name', 'surname', 'dob', 'subject', 'grade']
    if request.method == 'POST' and selected_db:
        db_path = get_db_path(user_id, selected_db)
        init_db(db_path)
        if 'excel_file' not in request.files:
            error = 'No file part.'
        else:
            file = request.files['excel_file']
            if file.filename == '':
                error = 'No selected file.'
            elif not file.filename.lower().endswith('.xlsx'):
                error = 'File must be .xlsx format.'
            else:
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                try:
                    wb = load_workbook(filepath)
                    ws = wb.active
                    headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    missing = [h for h in required_headers if h not in headers]
                    unsupported = [h for h in headers if h not in required_headers]
                    if missing:
                        error = f"Missing required headers: {', '.join(missing)}"
                    elif unsupported:
                        error = f"Unsupported headers found: {', '.join(unsupported)}"
                    else:
                        count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            data = dict(zip(headers, row))
                            # Data integrity checks
                            try:
                                forename = str(data['forename']).strip()
                                middle_name = str(data.get('middle_name', '')).strip()
                                surname = str(data['surname']).strip()
                                dob = str(data['dob']).strip()
                                subject = str(data['subject']).strip()
                                grade = str(data['grade']).strip()
                                if not forename or not surname or not dob or not subject or not grade:
                                    continue
                                grade_int = int(grade)
                                if not (0 <= grade_int <= 100):
                                    continue
                                import re
                                if not re.match(r"^\d{4}-\d{2}-\d{2}$", dob):
                                    continue
                            except Exception:
                                continue
                            conn = sqlite3.connect(db_path)
                            c = conn.cursor()
                            c.execute('''INSERT INTO person (forename, middle_name, surname, dob, subject, grade) VALUES (?, ?, ?, ?, ?, ?)''',
                                      (forename, middle_name, surname, dob, subject, grade_int))
                            conn.commit()
                            conn.close()
                            count += 1
                        success = f"Successfully imported {count} records."
                except Exception as e:
                    error = f"Error reading Excel file: {str(e)}"
                finally:
                    os.remove(filepath)
    return render_template('upload_excel.html', error=error, success=success, dbs=dbs, selected_db=selected_db)

if __name__ == '__main__':
    app.run(debug=True)
